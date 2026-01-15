import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
import os

# Define file path (save to ../examples/ by default or current dir)
output_file = os.path.join(os.path.dirname(__file__), "ISAAC_Metadata_Template_Final.xlsx")

# Workbook Setup
wb = Workbook()

# --- Common Styles ---
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
env_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid") # Green for Env
sample_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid") # Red for Sample

def style_header(ws, columns, row_idx=1):
    for col_num, (col_name, group) in enumerate(columns, 1):
        cell = ws.cell(row=row_idx, column=col_num, value=col_name)
        cell.font = header_font
        if group == 'Env':
            cell.fill = env_fill
        elif group == 'Sample':
            cell.fill = sample_fill
        else:
            cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 22

# ----------------------------
# 1. SHEET: Instrument Configs (The Beamline)
# ----------------------------
ws_config = wb.active
ws_config.title = "Instrument Configs"

config_columns_tuples = [
    ("Config ID", "Sys"),           
    ("Technique", "Sys"),           
    ("Mode", "Sys"),                
    ("Primary Signal Channel", "Sys"), 
    ("Normalization Channel", "Sys"),  
    ("Detectors", "Sys"),           
    ("Energy Calibration", "Sys"),  
    ("Monochromator", "Sys"),      # Added based on Angel's "Si(311)"
    ("Beam Energy (Nominal)", "Sys"), # Added based on Angel's Sheet 1
    ("Notes", "Sys")
]
style_header(ws_config, config_columns_tuples)

# Example Rows (Based on Angel's data)
ws_config.append(["Config_A", "XANES", "Fluorescence", "VortDT", "I0", "Canberra 100", "Pt-Foil 11564eV", "Si(311)", "11570", "Standard Setup"])
ws_config.append(["Config_B", "XAS", "Transmission", "I1", "I0", "Ion Chambers",  "Pt-Foil 11564eV", "Si(311)", "11570", ""])

# ----------------------------
# 2. SHEET: File List (The Science)
# ----------------------------
ws_files = wb.create_sheet("File List")

# Tuple: (Column Name, Group for Color Coding)
file_columns_tuples = [
    ("File Name", "Sys"), 
    ("Config ID", "Sys"),      # <--- LINKED to Instrument Configs
    
    # SAMPLE GROUP
    ("Sample Name", "Sample"), 
    ("Sample Formula", "Sample"), 
    ("Sample Form", "Sample"),
    ("Sample Description", "Sample"), # Angel's "Sample" column
    ("Sample Preparation", "Sample"), # Angel's "Working Electrode" details
    ("Loading (ug/cm2)", "Sample"),   # Angel's "Sample concentration"

    # ENVIRONMENT GROUP (Context)
    ("Environment", "Env"), 
    ("Reaction", "Env"), 
    
    ("Cell Type", "Env"),
    ("Counter Electrode", "Env"),     # Angel's Detail
    ("Reference Electrode", "Env"),   # Angel's Detail
    ("Flow Mode", "Env"),
    
    ("Solvent", "Env"),               # Angel's Detail
    ("Electrolyte Salt", "Env"),      # Angel's "Supporting electrolyte"
    ("pH", "Env"),
    
    ("Potential (V)", "Env"),
    ("Potential Scale", "Env"), 
    ("Gas Feed", "Env"),              # Angel's "Gas atmosphere"
    ("Pressure (atm)", "Env"),
    ("Temperature (C)", "Env")
]

style_header(ws_files, file_columns_tuples)

# ----------------------------
# 3. SHEET: Vocabulary (Hidden)
# ----------------------------
vocab_sheet = wb.create_sheet("Vocabulary")
vocab_sheet.sheet_state = 'hidden'

vocabs = {
    "Sample Form": ["film", "pellet", "powder", "monolith", "liquid", "slab_model", "gas"],
    "Environment": ["operando", "in_situ", "ex_situ", "in_silico"],
    "Cell Type": ["flow_cell", "h_cell", "beaker_cell", "gde_cell", "coin_cell", "three_electrode"],
    "Flow Mode": ["stagnant", "flowing_liquid", "bubbling_gas", "gas_diffusion", "rotating_disk"],
    "Potential Scale": ["RHE", "SHE", "Ag/AgCl", "SCE", "Hg/HgO", "V_vs_Ref"],
    "Reaction": ["CO2RR", "OER", "HER", "ORR", "None", "Pt Oxidation"]
}

# Write Vocabs
# Need a map of "Column Name" -> "Column Index"
col_map_files = {col[0]: i+1 for i, col in enumerate(file_columns_tuples)}

for i, (key, values) in enumerate(vocabs.items()):
    col_letter_vocab = chr(65 + i)
    vocab_sheet[f"{col_letter_vocab}1"] = key
    for r, val in enumerate(values):
        vocab_sheet[f"{col_letter_vocab}{r+2}"] = val
    
    vocab_range = f"Vocabulary!${col_letter_vocab}$2:${col_letter_vocab}${len(values)+1}"
    dv = DataValidation(type="list", formula1=vocab_range, allow_blank=True)
    dv.error = f"Please select a valid {key}"
    
    if key in col_map_files:
        target_col_idx = col_map_files[key]
        # logic for > 26 columns (AA, AB)
        if target_col_idx <= 26:
            col_letter = chr(64 + target_col_idx)
        else:
            # simple fix for 27th col 'AA' etc.
            # but we have approx 22 cols so A-Z is fine for now
            col_letter = chr(64 + target_col_idx)
            
        dv.add(f"{col_letter}2:{col_letter}1000")
        ws_files.add_data_validation(dv)

# ----------------------------
# 4. Special Validation: Config ID Link
# ----------------------------
config_id_range = "'Instrument Configs'!$A$2:$A$100"
dv_config = DataValidation(type="list", formula1=config_id_range, allow_blank=True)
dv_config.error = "Select a defined Instrument Config"
dv_config.add("B2:B1000")
ws_files.add_data_validation(dv_config)


# ----------------------------
# 5. SHEET: Campaign Info
# ----------------------------
info_sheet = wb.create_sheet("Campaign Info", 0)
info_sheet.sheet_view.showGridLines = False
info_sheet["A1"] = "ISAAC Campaign Metadata"
info_sheet["A1"].font = Font(size=14, bold=True)

fields = [
    ("Facility Name", "SSRL"),
    ("Organization", "SLAC"),
    ("Beamline/Instrument", "15-2"),
    ("Start Date", "2026-01-01")
]

for i, (field, default) in enumerate(fields):
    row = i + 3
    info_sheet[f"A{row}"] = field
    info_sheet[f"A{row}"].font = Font(bold=True)
    info_sheet[f"B{row}"] = default

instructions_start = len(fields) + 5
info_sheet[f"A{instructions_start}"] = "Instructions:"
info_sheet[f"A{instructions_start}"].font = Font(bold=True, underline="single")

lines = [
    "1. Define BEAMLINE settings in 'Instrument Configs' tab.",
    "2. List your files in 'File List' tab.",
    "3. Select the Equipment Config ID for each file.",
    "4. Fill in the green ENVIRONMENT columns (Cell, Electrodes, pH) for each file.",
    "5. Fill in the red SAMPLE columns (Prep, Loading) for each file."
]
for i, l in enumerate(lines):
    info_sheet[f"A{instructions_start + 1 + i}"] = l

# Save
wb.save(output_file)
print(f"Successfully created: {output_file}")
